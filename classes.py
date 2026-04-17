import numpy as np
from ultralytics import YOLO
import cv2
from PIL import Image
import requests
from transformers import AutoTokenizer, AutoProcessor, AutoModelForImageTextToText
from transformers import Qwen3VLForConditionalGeneration, AutoProcessor, Qwen2_5_VLForConditionalGeneration
from transformers import AutoModelForCausalLM, AutoTokenizer
import torch
import os
from transformers import LightOnOcrForConditionalGeneration, LightOnOcrProcessor
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side, Font, Alignment
from tqdm import tqdm

# --- Настройка глобального устройства ---
def get_best_device():
    if torch.cuda.is_available():
        return "cuda", torch.bfloat16 if torch.cuda.is_bf16_supported() else torch.float16
    elif torch.backends.mps.is_available():
        return "mps", torch.float32
    else:
        return "cpu", torch.float32

DEVICE, DTYPE = get_best_device()
print(f"Используется устройство: {DEVICE}, dtype: {DTYPE}")

class Parser():
    def __init__(self, model_path: str, OFFSET: float = 2, CONF_LEVEL: float = 0.5):
        # YOLO сам выберет GPU, если доступно. Явно укажем device.
        self.model = YOLO(model_path)
        self.model.to(DEVICE)  # Ultralytics поддерживает .to()
        self.OFFSET = OFFSET
        self.CONF_LEVEL = CONF_LEVEL

    def __resize_with_padding(self, image, target_size=(640, 640), background_color=(255, 255, 255)):
        h, w = image.shape[:2]
        scale = min(target_size[0] / h, target_size[1] / w)
        new_w, new_h = int(w * scale), int(h * scale)

        resized = cv2.resize(image, (new_w, new_h), interpolation=cv2.INTER_LANCZOS4)
        canvas = np.full((target_size[1], target_size[0], 3), background_color, dtype=np.uint8)

        x_offset = (target_size[0] - new_w) // 2
        y_offset = (target_size[1] - new_h) // 2
        canvas[y_offset:y_offset+new_h, x_offset:x_offset+new_w] = resized
        return canvas

    def predict(self, img_path: str, result_path: str):
        # YOLO predict уже использует GPU, если модель на GPU
        results = self.model.predict(source=img_path, conf=self.CONF_LEVEL, imgsz=640, device=DEVICE)
        for result in results:
            img = result.orig_img
            if result.masks is not None:
                for i, mask in enumerate(result.masks.data):
                    x1, y1, x2, y2 = map(int, result.boxes.xyxy[i])

                    x1_off = max(0, x1 - self.OFFSET)
                    y1_off = max(0, y1 - self.OFFSET)
                    x2_off = min(img.shape[1], x2 + self.OFFSET)
                    y2_off = min(img.shape[0], y2 + self.OFFSET)

                    cropped_img = img[y1_off:y2_off, x1_off:x2_off].copy()

                    full_mask = mask.cpu().numpy()
                    full_mask = cv2.resize(full_mask, (img.shape[1], img.shape[0]))

                    cropped_mask = full_mask[y1_off:y2_off, x1_off:x2_off]
                    binary_mask = (cropped_mask > 0.5).astype(np.uint8)

                    final_img = self.__resize_with_padding(cropped_img)
                    cv2.imwrite(f'{result_path}/{result.names[int(result.boxes.cls[i])]}.jpg', final_img)


class SimpleOCR():
    def __init__(self, save_directory: str = './local_model'):
        model_path = 'lightonai/LightOnOCR-2-1B-base'
        self.device = DEVICE
        self.dtype = DTYPE

        # Загружаем модель на GPU с поддержкой torch.compile (если версия позволяет)
        self.model = LightOnOcrForConditionalGeneration.from_pretrained(
            model_path,
            torch_dtype=self.dtype,
            device_map="auto"
        ).to(self.device)
        self.processor = LightOnOcrProcessor.from_pretrained(model_path)
        self.model.eval()
        # Включим torch.compile для ускорения инференса (опционально)
        if hasattr(torch, 'compile') and DEVICE == "cuda":
            try:
                self.model = torch.compile(self.model, mode="reduce-overhead")
            except Exception as e:
                print(f"torch.compile не поддерживается: {e}")

    @torch.inference_mode()
    def predict(self, img_path: str, max_new_tokens: int = 4096,
                prompt: str = 'extract the data from the image and provide it as a json file'):
        conversation = [{"role": "user", "content": [{"type": "image", "url": img_path}]}]
        inputs = self.processor.apply_chat_template(
            conversation,
            add_generation_prompt=True,
            tokenize=True,
            return_dict=True,
            return_tensors="pt",
        )
        # Явно переносим тензоры на GPU
        inputs = {k: v.to(device=self.device, dtype=self.dtype if v.is_floating_point() else None)
                  for k, v in inputs.items()}

        output_ids = self.model.generate(**inputs, max_new_tokens=1024)
        generated_ids = output_ids[0, inputs["input_ids"].shape[1]:]
        output_text = self.processor.decode(generated_ids, skip_special_tokens=True)
        return output_text


class HardOCR():
    def predict(self, img_path: str, output_type: str = 'flat-json'):
        url = "https://extraction-api.nanonets.com/extract"
        headers = {"Authorization": "Bearer c6163645-e3d4-11f0-9665-a23842209c4a"}
        files = {"file": open(img_path, "rb")}
        data = {"output_type": output_type}
        data["model"] = "nanonets"
        response = requests.post(url, headers=headers, files=files, data=data)
        return response.json()


class LLM():
    def __init__(self, model_name: str = "Qwen/Qwen3-1.7B"):
        self.tokenizer = AutoTokenizer.from_pretrained(model_name)
        self.model = AutoModelForCausalLM.from_pretrained(
            model_name,
            torch_dtype=DTYPE,
            device_map="auto"
        )
        self.model.eval()
        if hasattr(torch, 'compile') and DEVICE == "cuda":
            try:
                self.model = torch.compile(self.model, mode="reduce-overhead")
            except Exception as e:
                print(f"torch.compile не поддерживается: {e}")

    @torch.inference_mode()
    def predict(self, std_prompt: str = "извлеки мне из текста: '$$' значение всех колонок и полей...",
                list_content: list = []):
        prompt = std_prompt.replace('$$', '\n'.join(list_content))
        messages = [{"role": "user", "content": prompt}]
        text = self.tokenizer.apply_chat_template(
            messages,
            tokenize=False,
            add_generation_prompt=True,
            enable_thinking=True
        )
        model_inputs = self.tokenizer([text], return_tensors="pt").to(self.model.device)

        generated_ids = self.model.generate(**model_inputs, max_new_tokens=32768)
        output_ids = generated_ids[0][len(model_inputs.input_ids[0]):].tolist()

        try:
            index = len(output_ids) - output_ids[::-1].index(151668)
        except ValueError:
            index = 0

        thinking_content = self.tokenizer.decode(output_ids[:index], skip_special_tokens=True).strip("\n")
        content = self.tokenizer.decode(output_ids[index:], skip_special_tokens=True).strip("\n")
        return content.replace("```", '').replace('json', '').replace('\n', '').lower()


class VLLM():
    def __init__(self):
        model_path = 'Qwen/Qwen3-VL-2B-Instruct'
        self.model = Qwen3VLForConditionalGeneration.from_pretrained(
            model_path,
            torch_dtype=DTYPE,
            device_map="auto"
        )
        self.processor = AutoProcessor.from_pretrained(model_path)
        self.model.eval()
        if hasattr(torch, 'compile') and DEVICE == "cuda":
            try:
                self.model = torch.compile(self.model, mode="reduce-overhead")
            except Exception as e:
                print(f"torch.compile не поддерживается: {e}")

    @torch.inference_mode()
    def predict(self, img_path, prompt):
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "image", "image": img_path},
                    {"type": "text", "text": prompt},
                ],
            }
        ]
        inputs = self.processor.apply_chat_template(
            messages,
            tokenize=True,
            add_generation_prompt=True,
            return_dict=True,
            return_tensors="pt"
        )
        inputs = inputs.to(self.model.device)

        generated_ids = self.model.generate(**inputs, max_new_tokens=512)
        generated_ids_trimmed = [
            out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        output_text = self.processor.batch_decode(
            generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
        )
        return output_text[0].replace("```", '').replace('json', '').replace('\n', '').lower()


class Img2Json():
    def __init__(self, parser_path, parser_row_path, img_dir='./results'):
        self.parser = Parser(parser_path)
        self.parser_rows = Parser(parser_row_path, CONF_LEVEL=0.1)
        self.simple = SimpleOCR()
        self.hard = HardOCR()
        self.vllm = VLLM()
        self.llm = LLM()
        self.img_dir = img_dir

    def pars(self, img_path):
        lst = []
        prompt_simple='''
        извлеки мне из текста: '$$' следующие значения:
        \n сколько сдаю рулонов (в штуках)
        \n сколько шпуль (в штуках)
        \n дата
        \n номер заказа
        \n задание на перемотку
        \n сколько перемотать джамб (в штуках)
        \n пленки упак (в метрах)
        \n фамилия ответственного
        \n общий вес (в килограммах)
        \n номер смены
        \n продолжил задание (фамилия)
        \n задание закончил (фамилия)
        \n задание начал (фамилия)
        \n задание передал (фамилия)
        \n заголовок
        \nпредоставь в формате json, продоставь только значение этого файла, если какое-то значение пустое или только из нижних подчеркиваний то замени его на значение "nan" Очень важно именно в кавычках.
        '''
        prompt_vllm='''
        распарси фото-таблицу и достань от туда значения следующих ячеек:
        \n количество джамбов
        \n толщина, мкм
        \n ширина мм
        \n длина, м
        \n вес расчетный 1 дж
        \n вес расчетный общий
        \n
        \n учти, что количество джамбво находится в самом верху не превосходит 50 и является натуральным числом, возможно находится чуть правее надписи"количество джамбов"
        \n вес расчетный общий находится правее всего, поэтому возможно поехал немного вниз, учти это
        \n предоставь в формате json, продоставь только значение этого файла, если какое-то значение пустое или только из нижних подчеркиваний то замени его на значение "nan"
        '''
        img_dir = self.img_dir
        os.makedirs(img_dir, exist_ok=True)
        for i in range(1, 6):
            os.makedirs(f'{img_dir}/roll{i}', exist_ok=True)
        os.makedirs(f'{img_dir}/weight_djamb', exist_ok=True)

        self.parser.predict(img_path, img_dir)
        js = {}
        lst = []

        for file in tqdm(list(filter(lambda x: x.count('.') > 0, os.listdir(img_dir)))):
            if file == 'djamb.jpg':
                s = self.vllm.predict(img_dir + '/' + file, prompt_vllm)
                s = s.replace('json', '').replace('```', '').replace('\n', '').lower()
                try:
                    js.update(eval(s))
                except:
                    print(f"Ошибка eval для {file}")
                continue
            if file == 'table1.jpg':
                res = self.hard.predict(img_dir + '/' + file)
                if res['success']:
                    s = res['content']
                else:
                    print('ERROR: hard model is crashed')
                    prompt='''
                    распарси фото-таблицу и достань от туда значения следующих ячеек:
                    \n делить по длинне (значение данной ячейки находится в самом верху таблицы, представляет натурально число и находится правее надписи 'делить по длине')
                    \n информацию по каждому рулону (Номер рулона находится в самой левой колонке, предоставь в виде списка значений в строке с названием этого рулона. Если значение пропущено, но поставь 'nan')
                    \n обрезь (аналогично рулонам, необходио предоставить в виде list)
                    \n полезный вес (самая нижняя ячейка)
                    \n
                    \n в таблице присутствуют ровно 5 рулонов, каждый из них необходимо обработать, в каждом списке рулона должно быть ровно 6 значений, в порядке обхода таблицы слево на право по соответствующей строке.
                    \n все значения в ячейках представляют из себя либо вещественное число, либо натуральное, либо пустое значение.
                    \n под колонкой 'Резка ширины' находятся названия рулонов, пусть это будет ключом в json, остальные ячейки этой страки - лист значений.
                    \n очень важно сохранить пустые ячейки пустыми
                    \n предоставь ответ в формате json
                    '''
                    s = self.vllm.predict(img_dir + '/' + file, prompt)
                prompt='''
                распарси текст в виде json файла и достань от туда значения следующих ячеек:
                \n делить по длинне (представляет натурально число)
                \n информацию по каждому рулону (Номер рулона находится в самой левой колонке, предоставь в виде списка значений в строке с названием этого рулона. Если значение пропущено, но поставь 'nan')
                \n обрезь (аналогично рулонам, необходио предоставить в виде list)
                \n полезный вес
                \n
                \n в таблице присутствуют ровно 5 рулонов, каждый из них необходимо обработать, в каждом списке рулона должно быть ровно 6 значений.
                \n все значения в ячейках представляют из себя либо вещественное число, либо натуральное, либо пустое значение.
                \n очень важно сохранить пустые ячейки пустыми
                \n предоставь ответ в формате json
                \n вот тебе названия ключей списком: [делить по длине, рулон 1, рулон 2, рулон 3, рулон 4, рулон 5, обрезь, полезный вес]
                \n полезный вес - это одно число
                \n пример твоего ответа: '{"делить по длине":"2", "рулон 1":["23", "234", "nan", "12.32", "432", "4"], "рулон 2":["nan", "nan", "3432", "12.32", "432", "4"], "рулон 3":["235", "24", "32", "12.32", "432", "5"], "рулон 4":["23", "234", "43", "nan", "432", "4"], "рулон 5":["23", "234", "nan", "12.32", "432", "4"], "обрезь":["23", "234", "nan", "12.32", "432", "4"], "Полезный вес":"143.2"}'
                \n сам текст для анализа:
                \n $$
                '''
                d = self.llm.predict(std_prompt=prompt, list_content=[s])
                try:
                    js.update(eval(d))
                except:
                    print(f"Ошибка eval для table1")
                continue
            if file == 'title.jpg':
                js['заголовок'] = self.simple.predict(img_dir + '/' + file)
                continue
            if file[:4] == 'roll':
                folder_path = f'{img_dir}/{file[:-4]}'
                self.parser_rows.predict(f'{img_dir}/{file}', folder_path)
                num_roll = int(file[4])
                lst2 = [''] * 17
                for file2 in os.listdir(folder_path):
                    num_row = int(file2[3:-4])
                    val = self.vllm.predict(f'{folder_path}/{file2}',
                                             prompt="return ONLY VALUES. delete another text, delete ',' and join in in string all number")
                    if num_row != 0:
                        if num_roll % 2 == 1:
                            try:
                                val = float(f'{val[-3:-1]}.{val[-1]}')
                            except:
                                print(f'Ошибка в распознавании {num_roll} столбца, {num_row} строки')
                        else:
                            try:
                                val = f'{val[:2]}.{val[3]}'
                            except:
                                print(f'Ошибка в распознавании {num_roll} столбца, {num_row} строки')
                    lst2[num_row] = val
                js[f'вес рулона {num_roll}'] = lst2
                continue
            if file == 'weight_djamb.jpg':
                folder_path = f'{img_dir}/weight_djamb'
                self.parser_rows.predict(f'{img_dir}/{file}', folder_path)
                lst2 = [''] * 17
                for file2 in os.listdir(folder_path):
                    num_row = int(file2[3:-4])
                    val = self.vllm.predict(f'{folder_path}/{file2}',
                                             prompt="return ONLY VALUES. delete another text, delete ',' and join in in string all number")
                    try:
                        val = f'{val[:3]}.{val[3]}'
                    except:
                        print(f'Ошибка в распознавании 0 столбца, {num_row} строки')
                    lst2[num_row] = val
                js['вес джамбы'] = lst2
                continue
            if file == 'weight_trash.jpg':
                continue
            try:
                lst.append(self.simple.predict(img_dir + '/' + file))
            except:
                print(f"ошибка обработки {file}")

        g = str(self.llm.predict(prompt_simple, lst)).replace('null', 'nan')
        try:
            js.update(eval(g))
        except:
            print(f'ошибка обработки файла {file}\n вывод\n{g}')
        return js
class MakeTable():
    def __init__(self):
        self.er_fill= PatternFill(start_color='ff6d6d',
                                end_color='ff6d6d',
                                fill_type='solid')
        pass
    def __set_outer_border__(self, ws, cell_range, style='thick'):
        rows = list(ws[cell_range])
        thick = Side(style=style)

        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                left = thick if j == 0 else cell.border.left
                right = thick if j == len(row) - 1 else cell.border.right
                top = thick if i == 0 else cell.border.top
                bottom = thick if i == len(rows) - 1 else cell.border.bottom

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
    def __is_num__(self, s):
        if s=='nan':
            return ''
        try:
            return str(float(s.replace(',', '.'))).replace('.', ',')
        except:
            try:
                return str(float(s)).replace('.', ',')
            except:
                return ''
    def __add_str__(self, ws, cell, col_name, js):
        try:
            if(ws[cell]!='nan'):
                ws[cell]=js[col_name]
        except:
            ws[cell].fill=self.er_fill
            return
    def __add_num__(self, ws, cell, name, js, styling=True):
        try:
            num=js[name]
        except:
            ws[cell].fill=self.er_fill
            return
        ws[cell]=self.__is_num__(num)
        if styling and self.__is_num__(num)!='':
            ws[cell].fill=PatternFill(start_color='CCFFCC',
                                end_color='CCFFCC',
                                fill_type='solid')
    def __add_num_solo__(self, ws, cell, num, styling=True):
        ws[cell]=self.__is_num__(num)
        if styling and self.__is_num__(num)!='':
            ws[cell].fill=PatternFill(start_color='CCFFCC',
                                end_color='CCFFCC',
                                fill_type='solid')
    def make(self, save_path, js):
        wb = openpyxl.Workbook()
        ws = wb.active
        self.__set_outer_border__(ws, 'B1:H44', 'thin')
        for row in ws['B1:H48']:
            for cell in row:
                cell.font=Font(name='Times New Roman', size=11)
        columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H']

        for col in columns:
            ws.column_dimensions[col].width = 17

        green_fill = PatternFill(start_color='CCFFCC', # Светло-зеленый (как в Excel)
                                end_color='CCFFCC',
                                fill_type='solid')
        side = Side(style='thin')
        std_bord=Border(side, side, side, side)
        bld_font=Font(name='Times New Roman', size=11, bold=True)
        ws['B1']=js['заголовок']
        ws['B1'].font=Font(name='Times New Roman', size=10)

        ws['B2']='Спецификация'
        ws['B2'].font=Font(name='Times New Roman', size=14)

        ws['F2']='количество джамбов'
        ws['F2'].font=bld_font
        ws['F2'].alignment=Alignment('center')

        self.__add_num__(ws, 'F3', 'количество джамбов', js)
        ws['F3'].fill=green_fill
        ws['F3'].border=std_bord

        ws['B4']='Джамба'
        ws['B4'].font=bld_font

        ws['C4']='толщина, мкм'
        ws['D4']='ширина мм'
        ws['E4']='длина, м'
        ws['F4']='вес расчетный 1 дж'
        ws['H4']='вес расчетный общий'
        ws['H5'].border=std_bord
        ws['H4'].alignment=Alignment('right')

        for row in ws['C5:F5']:
            for cell in row:
                cell.border=std_bord
        self.__add_num__(ws, 'C5', 'толщина, мкм', js)
        self.__add_num__(ws, 'D5', 'ширина мм', js)
        self.__add_num__(ws, 'E5', 'длина, м', js)
        self.__add_num__(ws, 'F5', 'вес расчетный 1 дж', js)
        self.__add_num__(ws, 'H5', 'вес расчетный общий', js)
        ws['C7']='делить по длине'
        ws['C7'].font=bld_font
        ws['C7'].alignment=Alignment('right')

        ws['E7'].border=std_bord
        self.__add_num__(ws, 'E7', 'делить по длине', js)

        ws['B8']='Резка ширины'
        ws['D8']='Ширина, мм '
        ws['E8']='Длина, м'
        ws['F8']='Расчетный вес одного рулона'
        ws['F8'].alignment=Alignment(wrap_text=True)
        ws['G8']='Общий вес расчетный'
        ws['G8'].alignment=Alignment(wrap_text=True)
        ws['H8']='Количество роликов/шпули'
        ws['H8'].alignment=Alignment(wrap_text=True)

        for i in range(1, 6):
            ws[f'B{i+8}']=f'Рулон {i}'
            try:
                lst=js[f'рулон {i}']
            except:
                continue
            for num, ch in enumerate(['C', 'D', 'E', 'F', 'G', 'H']):
                self.__add_num_solo__(ws, f'{ch}{i+8}', lst[num])
        ws['B14']='Обрезь'
        try:
            lst=js['обрезь']
        except:
            pass
        for num, ch in enumerate(['C', 'D', 'E', 'F', 'G', 'H']):
            self.__add_num_solo__(ws, f'{ch}14', lst[num])
        for row in ws['B8:H14']:
            for cell in row:
                cell.border=std_bord
        ws['B15']='Полезный вес'
        ws['B15'].font=bld_font
        self.__add_num__(ws, 'G15', 'полезный вес', js)
        ws['G15'].border=std_bord
        self.__set_outer_border__(ws, 'B1:H16')


        ws['C17']='Заказ №'
        ws['C17'].font=bld_font
        ws['D17'].font=bld_font
        self.__add_str__(ws, 'D17', 'номер заказа', js)


        ws['B19']='Задание на перемотку  №'
        ws['B19'].font=bld_font
        ws['D19'].font=bld_font
        self.__add_str__(ws, 'D19', 'номер задания на перемотку', js)

        ws['E19']='Дата:'
        ws['E19'].font=bld_font
        ws['F19'].font=bld_font
        self.__add_str__(ws, 'F19', 'дата', js)

        ws['G19']='Смена:№'
        ws['G19'].font=bld_font
        ws['H19'].font=bld_font
        self.__add_str__(ws, 'H19', 'номер смены', js)

        ws['B20']='Перемотать'
        ws['D20']='Джамб'
        self.__add_str__(ws, 'C20', 'сколько джамб перемотать', js)

        ws['D21']='указывать чистый вес  - без веса  шпули!!!'
        ws['D21'].font=bld_font

        for row in ws['B22:H44']:
            for cell in row:
                cell.border=std_bord
        self.__set_outer_border__(ws, 'B44:H44')
        ws['B22']='вес джамбы'
        ws['B22'].font=bld_font
        for num, col in enumerate(['C', 'D', 'E', 'F', 'G']):
            ws[f'{col}22']=f'Вес рулона {num+1}'
            ws[f'{col}22'].font=bld_font
            ws['H22']='Вес брака/обрези'
            ws['H22'].font=bld_font
            ws['H22'].alignment=Alignment(wrap_text=True)

        cols=['B', 'C', 'D', 'E', 'F', 'G', 'H']
        for i, name in enumerate(['вес джамбы']+[f'вес рулона {i}' for i in range(1, 6)]+['вес брака/обрези']):
            try:
                lst=js[name]
            except:
                continue
            for j in range(len(lst)):
                self.__add_num_solo__(ws, f'{cols[i]}{j+23}', lst[j])

        ws['A44']='Итог'
        ws['A44'].font=bld_font

        ws['B46']='Сдаю:'
        ws['B46'].font=bld_font
        self.__add_num__(ws, 'C46', 'сколько сдаю рулонов', js, False)
        ws['C46']=f'{str(ws["C46"].value)} рулонов'
        ws['D46'].font=bld_font

        ws['E46']='Общий вес:'
        self.__add_num__(ws, 'F46', 'общий вес в кг', js, False)
        ws['E46'].font=bld_font
        ws['F46']=f'{str(ws["F46"].value)} кг'

        ws['G46']='Брак:'
        self.__add_num__(ws, 'H46', 'брак в кг', js, False)
        ws['G46'].font=bld_font
        ws['H46']=f'{str(ws["H46"].value)} кг'


        ws['C48']='Ответственный:'
        ws['C48'].font=bld_font
        self.__add_str__(ws, 'D48', 'фамилия ответственного', js)
        wb.save(save_path)
class Img2Xlsx():
    def __init__(self, parser_path='./utils/parser.pt', pars_dir='./utils/pars_img'):
        self.img2js=Img2Json(parser_path, pars_dir)
        self.maketable=MakeTable()
    def convert(self, img_path, table_path):
        js=self.img2js.pars(img_path)
        self.maketable.make(table_path, js)